import tkinter as tk
from tkinter import ttk, messagebox, font
import openpyxl
from datetime import datetime
import os
import shutil
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

# Configuration
FICHIER = "stock.xlsx"
CLE_USB_PATH = "E:/BackupStock"

# Thèmes améliorés
THEME_CLAIR = {
    "bg": "#f5f7fa",
    "fg": "#2d3748",
    "button_bg": "#4f46e5",
    "button_fg": "#ffffff",
    "entry_bg": "#ffffff",
    "highlight": "#e2e8f0",
    "success": "#10b981",
    "danger": "#ef4444",
    "warning": "#f59e0b",
    "accent": "#6366f1",
    "card": "#ffffff",
    "text_muted": "#64748b"
}

THEME_SOMBRE = {
    "bg": "#1e293b",
    "fg": "#f8fafc",
    "button_bg": "#6366f1",
    "button_fg": "#ffffff",
    "entry_bg": "#334155",
    "highlight": "#475569",
    "success": "#10b981",
    "danger": "#ef4444",
    "warning": "#f59e0b",
    "accent": "#818cf8",
    "card": "#1e293b",
    "text_muted": "#94a3b8"
}

theme_actuel = THEME_CLAIR

# Icônes
ICONS = {
    "add": "📥 Ajouter",
    "sell": "📤 Vendre",
    "delete": "🗑️ Supprimer",
    "search": "🔍",
    "print": "🖨️",
    "clear": "🧹",
    "theme": "🌓",
    "stock": "📦",
    "in": "⬆️",
    "out": "⬇️"
}


def init_fichier():
    if not os.path.exists(FICHIER):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produits"
        ws.append(["Nom Produit", "Stock"])
        wb.create_sheet("Entrees").append(["Date", "Nom", "Quantité"])
        wb.create_sheet("Sorties").append(["Date", "Nom", "Quantité"])
        wb.save(FICHIER)


def sauvegarde_auto():
    if os.path.exists(CLE_USB_PATH):
        try:
            shutil.copy(FICHIER, CLE_USB_PATH)
            maj_status(f"Sauvegarde vers clé USB réussie.", "success")
        except Exception as e:
            maj_status(f"Erreur sauvegarde : {e}", "danger")
    else:
        maj_status("Clé USB non détectée, sauvegarde locale uniquement.", "warning")


def maj_status(texte, type="info"):
    colors = {
        "info": theme_actuel["fg"],
        "success": theme_actuel["success"],
        "danger": theme_actuel["danger"],
        "warning": theme_actuel["warning"]
    }
    status_label.config(text=texte, fg=colors.get(type, theme_actuel["fg"]))


def ajouter():
    nom_produit = entry_nom.get().strip()
    quantite = entry_qte.get().strip()

    if not nom_produit or not quantite.isdigit():
        messagebox.showwarning("Erreur", "Veuillez saisir un nom et une quantité valide.")
        return

    qte = int(quantite)
    wb = openpyxl.load_workbook(FICHIER)
    ws = wb["Produits"]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == nom_produit:
            row[1].value = (row[1].value or 0) + qte
            break
    else:
        ws.append([nom_produit, qte])

    wb["Entrees"].append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), nom_produit, qte])
    wb.save(FICHIER)
    sauvegarde_auto()
    afficher_stock()
    charger_historique()
    entry_nom.delete(0, tk.END)
    entry_qte.delete(0, tk.END)
    maj_status(f"Produit '{nom_produit}' ajouté avec succès.", "success")


def vendre():
    nom_produit = entry_nom.get().strip()
    quantite = entry_qte.get().strip()

    if not nom_produit or not quantite.isdigit():
        messagebox.showwarning("Erreur", "Veuillez saisir un nom et une quantité valide.")
        return

    qte = int(quantite)
    wb = openpyxl.load_workbook(FICHIER)
    ws = wb["Produits"]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == nom_produit:
            if (row[1].value or 0) >= qte:
                row[1].value -= qte
                wb["Sorties"].append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), nom_produit, qte])
                wb.save(FICHIER)
                sauvegarde_auto()
                afficher_stock()
                charger_historique()
                entry_nom.delete(0, tk.END)
                entry_qte.delete(0, tk.END)
                maj_status(f"Vente de '{nom_produit}' enregistrée.", "success")
                return
            else:
                messagebox.showerror("Erreur", "Stock insuffisant.")
                return
    messagebox.showerror("Erreur", "Produit non trouvé.")


def supprimer():
    nom_produit = entry_nom.get().strip()
    if not nom_produit:
        messagebox.showwarning("Erreur", "Veuillez saisir un nom de produit.")
        return

    if not messagebox.askyesno("Confirmation", f"Voulez-vous vraiment supprimer '{nom_produit}' ?"):
        return

    wb = openpyxl.load_workbook(FICHIER)
    ws = wb["Produits"]

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == nom_produit:
            ws.delete_rows(i)
            wb.save(FICHIER)
            sauvegarde_auto()
            afficher_stock()
            charger_historique()
            entry_nom.delete(0, tk.END)
            entry_qte.delete(0, tk.END)
            maj_status(f"Produit '{nom_produit}' supprimé.", "success")
            return
    messagebox.showerror("Erreur", "Produit non trouvé.")


def afficher_stock(filtre=""):
    wb = openpyxl.load_workbook(FICHIER)
    ws = wb["Produits"]
    listbox_stock.delete(0, tk.END)

    produits = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if filtre.lower() in str(row[0]).lower():
            produits.append(row)

    produits.sort(key=lambda x: str(x[0]).lower())

    for row in produits:
        stock_text = f"{row[0]:<30} | Stock: {row[1] if row[1] is not None else 0:>5}"
        listbox_stock.insert(tk.END, stock_text)

        if (row[1] or 0) < 5:
            listbox_stock.itemconfig(tk.END, {'fg': theme_actuel["danger"]})


def charger_historique():
    wb = openpyxl.load_workbook(FICHIER)

    # Historique des entrées
    listbox_entree.delete(0, tk.END)
    entrees = []
    for row in wb["Entrees"].iter_rows(min_row=2, values_only=True):
        entrees.append(row)

    entrees.sort(key=lambda x: datetime.strptime(x[0], "%Y-%m-%d %H:%M:%S"), reverse=True)

    for row in entrees:
        listbox_entree.insert(tk.END, f"{row[0]:<20} | {row[1]:<30} | +{row[2]:>5}")
        listbox_entree.itemconfig(tk.END, {'fg': theme_actuel["success"]})

    # Historique des sorties
    listbox_sortie.delete(0, tk.END)
    sorties = []
    for row in wb["Sorties"].iter_rows(min_row=2, values_only=True):
        sorties.append(row)

    sorties.sort(key=lambda x: datetime.strptime(x[0], "%Y-%m-%d %H:%M:%S"), reverse=True)

    for row in sorties:
        listbox_sortie.insert(tk.END, f"{row[0]:<20} | {row[1]:<30} | -{row[2]:>5}")
        listbox_sortie.itemconfig(tk.END, {'fg': theme_actuel["danger"]})


def exporter_pdf():
    wb = openpyxl.load_workbook(FICHIER)
    ws = wb["Produits"]
    c = canvas.Canvas("stock_export.pdf", pagesize=A4)
    width, height = A4

    c.setFillColor(theme_actuel["button_bg"])
    c.rect(0, height - 80, width, 80, fill=True, stroke=False)
    c.setFont("Helvetica-Bold", 24)
    c.setFillColorRGB(1, 1, 1)
    c.drawString(50, height - 60, "Inventaire des Stocks")
    c.setFont("Helvetica", 10)
    c.drawString(50, height - 80, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    y = height - 120
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(theme_actuel["fg"])
    c.drawString(50, y, "Produit")
    c.drawString(width - 150, y, "Stock")
    y -= 20

    c.setFont("Helvetica", 12)
    for row in ws.iter_rows(min_row=2, values_only=True):
        c.drawString(50, y, row[0])
        c.drawString(width - 150, y, str(row[1] if row[1] is not None else 0))
        y -= 20
        if y < 50:
            c.showPage()
            y = height - 50
            c.setFont("Helvetica", 12)

    c.save()
    maj_status("PDF exporté avec succès.", "success")


def imprimer_stock():
    try:
        print_window = tk.Toplevel(root)
        print_window.title("Aperçu avant impression")

        text = tk.Text(print_window, font=('Courier New', 10))
        text.pack(fill='both', expand=True)

        # Titre et date
        text.insert('end', "RAPPORT COMPLET DE STOCK\n", 'header')
        text.insert('end', f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")

        # En-têtes des colonnes
        text.insert('end', f"{'PRODUIT':<30} {'ENTREES':>10} {'SORTIES':>10} {'STOCK':>10}\n", 'header')
        text.insert('end', "-" * 62 + "\n")

        wb = openpyxl.load_workbook(FICHIER)
        ws_produits = wb["Produits"]
        ws_entrees = wb["Entrees"]
        ws_sorties = wb["Sorties"]

        # Dictionnaire pour stocker les totaux
        produits = {}

        # Récupérer tous les produits
        for row in ws_produits.iter_rows(min_row=2, values_only=True):
            nom = row[0]
            if nom:
                produits[nom] = {
                    'stock': row[1] if row[1] is not None else 0,
                    'entrees': 0,
                    'sorties': 0
                }

        # Calculer les entrées
        for row in ws_entrees.iter_rows(min_row=2, values_only=True):
            nom = row[1]
            qte = row[2] if row[2] is not None else 0
            if nom in produits:
                produits[nom]['entrees'] += qte

        # Calculer les sorties
        for row in ws_sorties.iter_rows(min_row=2, values_only=True):
            nom = row[1]
            qte = row[2] if row[2] is not None else 0
            if nom in produits:
                produits[nom]['sorties'] += qte

        # Trier les produits par nom
        produits_tries = sorted(produits.items(), key=lambda x: x[0].lower())

        # Afficher chaque produit avec ses statistiques
        for nom, data in produits_tries:
            text.insert('end',
                        f"{nom:<30} {data['entrees']:>10} {data['sorties']:>10} {data['stock']:>10}\n")

            # Colorer les stocks faibles en rouge
            if data['stock'] < 5:
                text.tag_add('low_stock', f'{float(text.index("end")) - 1} linestart',
                             f'{float(text.index("end")) - 1} lineend')

        # Style pour l'en-tête et les stocks faibles
        text.tag_configure('header', font=('Courier New', 10, 'bold'))
        text.tag_configure('low_stock', foreground='red')

        def print_text():
            try:
                import tempfile
                import os

                fd, temp_path = tempfile.mkstemp(suffix='.txt')
                with os.fdopen(fd, 'w') as f:
                    f.write(text.get('1.0', 'end'))

                if os.name == 'nt':
                    os.startfile(temp_path, 'print')
                elif os.name == 'posix':
                    if os.uname().sysname == 'Darwin':
                        os.system(f'lpr {temp_path}')
                    else:
                        os.system(f'lp {temp_path}')

                print_window.after(5000, lambda: os.unlink(temp_path))
                maj_status("Impression envoyée à l'imprimante", "success")
                print_window.destroy()
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible d'imprimer : {str(e)}")
                maj_status(f"Erreur d'impression : {str(e)}", "danger")

        print_button = tk.Button(print_window, text="Imprimer", command=print_text)
        print_button.pack(pady=10)

    except Exception as e:
        messagebox.showerror("Erreur", f"Impossible de préparer l'impression : {str(e)}")
        maj_status(f"Erreur d'impression : {str(e)}", "danger")


def vider_historique(nom_feuille):
    if not messagebox.askyesno("Confirmation", f"Voulez-vous vraiment vider l'historique {nom_feuille} ?"):
        return

    wb = openpyxl.load_workbook(FICHIER)
    ws = wb[nom_feuille]
    ws.delete_rows(2, ws.max_row)
    wb.save(FICHIER)
    charger_historique()
    maj_status(f"Historique {nom_feuille} vidé.", "success")


def rechercher():
    filtre = entry_recherche.get().strip()
    afficher_stock(filtre)


def basculer_theme():
    global theme_actuel
    theme_actuel = THEME_SOMBRE if theme_actuel == THEME_CLAIR else THEME_CLAIR
    appliquer_theme()
    afficher_stock()
    charger_historique()


def appliquer_theme():
    style = ttk.Style()
    style.theme_use('clam')

    root.config(bg=theme_actuel["bg"])

    for widget in [label_nom, label_qte, label_recherche, status_label]:
        widget.config(bg=theme_actuel["bg"], fg=theme_actuel["fg"])

    for entry in [entry_nom, entry_qte, entry_recherche]:
        entry.config(
            bg=theme_actuel["entry_bg"],
            fg=theme_actuel["fg"],
            insertbackground=theme_actuel["fg"],
            highlightcolor=theme_actuel["accent"],
            highlightthickness=1,
            relief="flat"
        )

    for listbox in [listbox_stock, listbox_entree, listbox_sortie]:
        listbox.config(
            bg=theme_actuel["entry_bg"],
            fg=theme_actuel["fg"],
            selectbackground=theme_actuel["accent"],
            selectforeground="white",
            font=('Consolas', 10),
            relief="flat"
        )

    button_config = {
        "font": ('Helvetica', 10),
        "borderwidth": 0,
        "relief": "flat",
        "padx": 10,
        "pady": 5
    }

    button_ajouter.config(bg=theme_actuel["success"], fg="white", **button_config)
    button_vendre.config(bg=theme_actuel["danger"], fg="white", **button_config)
    button_supprimer.config(bg=theme_actuel["warning"], fg="white", **button_config)
    button_rechercher.config(bg=theme_actuel["accent"], fg="white", **button_config)
    button_pdf.config(bg=theme_actuel["accent"], fg="white", **button_config)
    button_imprimer.config(bg=theme_actuel["accent"], fg="white", **button_config)
    button_vider_entree.config(bg=theme_actuel["warning"], fg="white", **button_config)
    button_vider_sortie.config(bg=theme_actuel["warning"], fg="white", **button_config)
    button_theme.config(bg=theme_actuel["highlight"], fg=theme_actuel["fg"], **button_config)

    style.configure('TNotebook', background=theme_actuel["bg"], borderwidth=0)
    style.configure('TNotebook.Tab',
                    background=theme_actuel["highlight"],
                    foreground=theme_actuel["fg"],
                    padding=[10, 5],
                    font=('Helvetica', 10))
    style.map('TNotebook.Tab',
              background=[('selected', theme_actuel["accent"])],
              foreground=[('selected', 'white')])

    for frame in [input_frame, search_frame, bottom_frame, status_frame]:
        frame.config(bg=theme_actuel["bg"])


# Initialisation
init_fichier()
root = tk.Tk()
root.title("📦 GESTION DE STOCK")
root.geometry("1000x700")
root.minsize(900, 600)

# Police
custom_font = font.Font(family="Helvetica", size=10)

# Cadre principal
main_frame = tk.Frame(root, bg=theme_actuel["bg"], padx=10, pady=10)
main_frame.pack(fill="both", expand=True)

# En-tête
header_frame = tk.Frame(main_frame, bg=theme_actuel["bg"])
header_frame.pack(fill="x")

title_label = tk.Label(header_frame, text="📦 GESTION DE STOCK",
                       bg=theme_actuel["bg"], fg=theme_actuel["accent"],
                       font=('Helvetica', 16, 'bold'))
title_label.pack(side="left")

button_theme = tk.Button(header_frame, text=ICONS["theme"], command=basculer_theme,
                         font=custom_font, bg=theme_actuel["highlight"],
                         fg=theme_actuel["fg"], borderwidth=0)
button_theme.pack(side="right")

# Section de saisie
input_frame = tk.LabelFrame(main_frame, text=" Gestion des Produits ",
                            bg=theme_actuel["bg"], fg=theme_actuel["fg"],
                            font=('Helvetica', 12, 'bold'), padx=10, pady=10)
input_frame.pack(fill="x", pady=(0, 10))

label_nom = tk.Label(input_frame, text="Nom du Produit :",
                     bg=theme_actuel["bg"], fg=theme_actuel["fg"],
                     font=custom_font)
label_nom.grid(row=0, column=0, padx=5, pady=5, sticky="w")

entry_nom = tk.Entry(input_frame, font=custom_font, width=30)
entry_nom.grid(row=0, column=1, padx=5, pady=5)

label_qte = tk.Label(input_frame, text="Quantité :",
                     bg=theme_actuel["bg"], fg=theme_actuel["fg"],
                     font=custom_font)
label_qte.grid(row=1, column=0, padx=5, pady=5, sticky="w")

entry_qte = tk.Entry(input_frame, font=custom_font, width=10)
entry_qte.grid(row=1, column=1, padx=5, pady=5, sticky="w")

# Boutons d'action
button_frame = tk.Frame(input_frame, bg=theme_actuel["bg"])
button_frame.grid(row=0, column=2, rowspan=2, padx=20)

button_ajouter = tk.Button(button_frame, text=ICONS["add"], command=ajouter,
                           font=custom_font, width=15)
button_ajouter.grid(row=0, column=0, padx=5, pady=5)

button_vendre = tk.Button(button_frame, text=ICONS["sell"], command=vendre,
                          font=custom_font, width=15)
button_vendre.grid(row=0, column=1, padx=5, pady=5)

button_supprimer = tk.Button(button_frame, text=ICONS["delete"], command=supprimer,
                             font=custom_font, width=15)
button_supprimer.grid(row=0, column=2, padx=5, pady=5)

# Recherche
search_frame = tk.Frame(main_frame, bg=theme_actuel["bg"])
search_frame.pack(fill="x", pady=(0, 10))

label_recherche = tk.Label(search_frame, text=f"{ICONS['search']} Rechercher :",
                           bg=theme_actuel["bg"], fg=theme_actuel["fg"],
                           font=custom_font)
label_recherche.pack(side="left", padx=(0, 5))

entry_recherche = tk.Entry(search_frame, font=custom_font, width=40)
entry_recherche.pack(side="left", padx=5, pady=5, fill="x", expand=True)

button_rechercher = tk.Button(search_frame, text="Rechercher", command=rechercher,
                              font=custom_font)
button_rechercher.pack(side="left", padx=(5, 0))

# Onglets
notebook = ttk.Notebook(main_frame)
notebook.pack(fill="both", expand=True)

# Onglet Stock
frame_stock = tk.Frame(notebook, bg=theme_actuel["bg"])
scroll_stock = tk.Scrollbar(frame_stock)
listbox_stock = tk.Listbox(frame_stock, yscrollcommand=scroll_stock.set, width=100, height=20)
scroll_stock.config(command=listbox_stock.yview)

scroll_stock.pack(side="right", fill="y")
listbox_stock.pack(side="left", fill="both", expand=True)

# Onglet Entrées
frame_entree = tk.Frame(notebook, bg=theme_actuel["bg"])
scroll_entree = tk.Scrollbar(frame_entree)
listbox_entree = tk.Listbox(frame_entree, yscrollcommand=scroll_entree.set, width=100, height=20)
scroll_entree.config(command=listbox_entree.yview)

scroll_entree.pack(side="right", fill="y")
listbox_entree.pack(side="left", fill="both", expand=True)

# Onglet Sorties
frame_sortie = tk.Frame(notebook, bg=theme_actuel["bg"])
scroll_sortie = tk.Scrollbar(frame_sortie)
listbox_sortie = tk.Listbox(frame_sortie, yscrollcommand=scroll_sortie.set, width=100, height=20)
scroll_sortie.config(command=listbox_sortie.yview)

scroll_sortie.pack(side="right", fill="y")
listbox_sortie.pack(side="left", fill="both", expand=True)

notebook.add(frame_stock, text=f" {ICONS['stock']} Stock Actuel ")
notebook.add(frame_entree, text=f" {ICONS['in']} Entrées ")
notebook.add(frame_sortie, text=f" {ICONS['out']} Sorties ")

# Barre d'outils
bottom_frame = tk.Frame(main_frame, bg=theme_actuel["bg"])
bottom_frame.pack(fill="x", pady=(10, 0))

button_pdf = tk.Button(bottom_frame, text=f"{ICONS['print']} PDF", command=exporter_pdf,
                       font=custom_font)
button_pdf.pack(side="left", padx=5, pady=5)

button_imprimer = tk.Button(bottom_frame, text=f"{ICONS['print']} Imprimer",
                            command=imprimer_stock, font=custom_font)
button_imprimer.pack(side="left", padx=5, pady=5)

button_vider_entree = tk.Button(bottom_frame, text=f"{ICONS['clear']} Entrées",
                                command=lambda: vider_historique("Entrees"),
                                font=custom_font)
button_vider_entree.pack(side="left", padx=5, pady=5)

button_vider_sortie = tk.Button(bottom_frame, text=f"{ICONS['clear']} Sorties",
                                command=lambda: vider_historique("Sorties"),
                                font=custom_font)
button_vider_sortie.pack(side="left", padx=5, pady=5)

# Barre de statut
status_frame = tk.Frame(main_frame, bg=theme_actuel["highlight"], height=25)
status_frame.pack(fill="x", pady=(10, 0))

status_label = tk.Label(status_frame, text="Prêt.", anchor="w",
                        bg=theme_actuel["highlight"], fg=theme_actuel["fg"])
status_label.pack(fill="x", padx=10)

# Initialisation
appliquer_theme()
afficher_stock()
charger_historique()

# Focus
entry_nom.focus_set()

root.mainloop()
